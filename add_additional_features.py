from openpyxl import load_workbook

wb = load_workbook('features cleaned.xlsx')
ws = wb['Sheet1']

with open('snr_values.txt') as f:
	snr = f.read()

with open('hnr_values.txt') as f:
	hnr = f.read()

snr = snr.split(',')

hnr = hnr.split(',')

ws['CL1'] = 'snr_values'
ws['CM1'] = 'hnr_values'
for idx, item in enumerate(snr):
	
	index = 'CL' + str(idx + 2)
	ws[index] = item

for idx, item in enumerate(hnr):

	index = 'CM' + str(idx + 2)
	ws[index] = item

diagnosis = 0
ws['CN1'] = 'Diagnosis'

for i in range(1, 209):
	if i < 10:
		with open('./physionet.org/files/voiced/1.0.0/voice00' + str(i) + '-info.txt') as f:
			lines = f.readlines()
	elif i < 100:
		with open('./physionet.org/files/voiced/1.0.0/voice0' + str(i) + '-info.txt') as f:
			lines = f.readlines()
	else:
		with open('./physionet.org/files/voiced/1.0.0/voice' + str(i) + '-info.txt') as f:
			lines = f.readlines()

	line = lines[4]
	line = line.split()
	diagnosis = ' '.join(line[1:])

	index = 'CN' + str(i + 1)
	ws[index] = diagnosis

wb.save('features cleaned.xlsx')